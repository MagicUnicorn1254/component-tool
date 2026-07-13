import streamlit as st
import io
import re
import requests
import logging
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

logging.getLogger("pypdf").setLevel(logging.CRITICAL)

st.set_page_config(page_title="Component Engine", layout="wide")

CLIENT_ID = '295EGpEwJEuPCaTsslUztQdBUXQOCLvGztU2UlEkqGfcIyur'
CLIENT_SECRET = 'X80tbzNKh50mx6IieAOoJcWl57jhE3dmiycn2jYj74XTVZisrUGyJHKumFiB1wDr'

CORE_PDF_TARGETS = {
    "power": ["power dissipation", "power consumption", "p_diss", "max. power", "power"],
    "voltage": ["supply voltage", "voltage - supply", "operating voltage", "rated voltage", "voltage"],
    "current": ["continuous drain current", "drain current", "id", "supply current", "operating current", "current"],
    "temperature": ["operating temperature", "storage temperature", "temp range", "case temperature"],
    "min temperature": ["min temperature", "minimum temperature", "min temp"], 
    "max temperature": ["max temperature", "maximum temperature", "max temp"], 
    "wavelength": ["wavelength", "wave length", "optical wavelength"],
    "data rate": ["data rate", "speed"],
    "vds": ["drain to source voltage", "drain-source voltage", "vds", "vdss"],
    "vgs": ["gate to source voltage", "gate-source voltage", "vgs", "vgss"],
    "rds on": ["rds(on)", "drain-source on-state resistance", "on-resistance", "rds on (max)"],
    "gate charge": ["gate charge", "qg"],
    "forward voltage": ["voltage - forward", "forward voltage", "vf"],
    "luminous intensity": ["millicandela rating", "luminous intensity", "brightness", "mcd"],
    "viewing angle": ["viewing angle", "beam angle"],
    "color": ["emitted color", "color"],
    "sound level": ["sound pressure level", "spl", "db"]
}

@st.cache_data(ttl=86400) 
def get_usd_to_cad_rate():
    try:
        url = "https://open.er-api.com/v6/latest/USD"
        response = requests.get(url, timeout=3)
        return float(response.json()['rates']['CAD'])
    except Exception:
        return 1.36 

def get_digikey_token():
    url = "https://api.digikey.com/v1/oauth2/token"
    payload = {'grant_type': 'client_credentials', 'client_id': CLIENT_ID, 'client_secret': CLIENT_SECRET}
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    response = requests.post(url, data=payload, headers=headers)
    if response.status_code != 200: 
        st.error("Authentication failed! Check credentials.")
        return None
    return response.json()['access_token']

def extract_prices(prod_dict):
    """Pulls both the 1-Unit Price and 100-Unit Price"""
    price_1 = 0.0
    price_100 = 0.0
    try:
        variations = prod_dict.get('ProductVariations', [])
        if not variations:
            variations = [{'StandardPricing': prod_dict.get('StandardPricing', [])}]
            
        best_1 = float('inf')
        best_100 = float('inf')
        
        for var in variations:
            pricing_list = var.get('StandardPricing', [])
            if not pricing_list: continue
                
            sorted_breaks = sorted(pricing_list, key=lambda x: x.get('BreakQuantity', 0))
            
            for pb in sorted_breaks:
                if pb.get('BreakQuantity', 0) == 1:
                    if pb.get('UnitPrice', 0.0) < best_1:
                        best_1 = pb.get('UnitPrice', 0.0)
                        
            if sorted_breaks[0].get('BreakQuantity', 0) <= 100:
                current_var_100 = float('inf')
                for pb in sorted_breaks:
                    if pb.get('BreakQuantity', 0) <= 100:
                        current_var_100 = pb.get('UnitPrice', 0.0)
                if 0.0 < current_var_100 < best_100:
                    best_100 = current_var_100
                    
        if best_1 == float('inf'):
            all_prices = [p.get('UnitPrice', 0.0) for v in variations for p in v.get('StandardPricing', [])]
            if all_prices: best_1 = max(all_prices)
            else: best_1 = 0.0
            
        if best_100 != float('inf'): price_100 = best_100 * 100
            
        return best_1, price_100
    except Exception: 
        return 0.0, 0.0

def split_temperature_ranges(part_dict):
    for k, v in list(part_dict.items()):
        if "temperature" in k and "~" in str(v):
            match = re.search(r"([-+]?\d+(?:\.\d+)?)[^0-9\-+]*~[^0-9\-+]*([-+]?\d+(?:\.\d+)?)", str(v))
            if match:
                part_dict['min temperature'] = match.group(1) + "°C"
                part_dict['max temperature'] = match.group(2) + "°C"
            break

def auto_extract_specs_from_pdf(pdf_url, expected_keys):
    if not pdf_url: return {}
    if pdf_url.startswith("//"): pdf_url = "https:" + pdf_url
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(pdf_url, headers=headers, timeout=5)
        if response.status_code != 200: return {}
        
        reader = PdfReader(io.BytesIO(response.content))
        max_pages = min(3, len(reader.pages))
        full_text = "".join(reader.pages[i].extract_text() or "" for i in range(max_pages))
        
        extracted_params = {}
        for category, aliases in CORE_PDF_TARGETS.items():
            pattern = re.compile(rf"({'|'.join(sorted(aliases, key=len, reverse=True))})[^\d]*([0-9\.\-\~]+\s*[a-zA-Z]*)", re.IGNORECASE)
            match = pattern.search(full_text)
            if match: extracted_params[category] = match.group(2).split('\n')[0].strip()

        for key in expected_keys:
            if key in extracted_params: continue
            search_terms = [key]
            acronyms = re.findall(r'\(([^)]+)\)', key)
            search_terms.extend(acronyms)
            if "-" in key:
                search_terms.append(key.split("-")[0].strip())
            search_terms = [re.escape(t) for t in search_terms if len(t) > 1]
            if not search_terms: continue
                
            pattern_str = '|'.join(sorted(search_terms, key=len, reverse=True))
            pattern = re.compile(rf"({pattern_str})[^\d]*([0-9\.\-\~]+\s*[a-zA-Z]*)", re.IGNORECASE)
            
            match = pattern.search(full_text)
            if match: 
                extracted_params[key] = match.group(2).split('\n')[0].strip()
                
        split_temperature_ranges(extracted_params)
        return extracted_params
    except Exception: 
        return {}

def fetch_part_data(token, part_number):
    part_number = part_number.strip()
    url = "https://api.digikey.com/products/v4/search/keyword"
    headers = {
        "Authorization": f"Bearer {token}", 
        "X-DIGIKEY-Client-Id": CLIENT_ID, 
        "Content-Type": "application/json",
        "X-DIGIKEY-Locale-Site": "US", 
        "X-DIGIKEY-Locale-Currency": "CAD"
    }
    
    response = requests.post(url, json={"Keywords": part_number, "Limit": 10}, headers=headers)
    if response.status_code != 200 or not response.json().get('Products'):
        return None 

    products = response.json()['Products']
    prod = None
    for p in products:
        actual_part_number = p.get('ManufacturerProductNumber') or p.get('ProductCode') or ""
        # Accepts standard Part Number OR hidden DigiKey Barcodes
        if actual_part_number.lower().startswith(part_number.lower()) or part_number.lower() in str(p).lower():
            prod = p
            break
            
    if not prod:
        raise Exception(f"Product '{part_number}' does not exist. Please check your part number.")

    raw_price_1, raw_price_100 = extract_prices(prod)
    is_usd = (headers.get("X-DIGIKEY-Locale-Site") == "US")
    cad_exchange_rate = get_usd_to_cad_rate() if is_usd else 1.0

    datasheet_url = prod.get('DatasheetUrl')

    part_profile = {
        "Part Number": prod.get('ManufacturerProductNumber') or prod.get('ProductCode') or part_number,
        "Manufacturer": prod.get('Manufacturer', {}).get('Name'),
        "Description": prod.get('ProductDescription') or "",
        "Stock": prod.get('QuantityAvailable', 0),
        "Price1": raw_price_1 * cad_exchange_rate,
        "Price100": raw_price_100 * cad_exchange_rate,
        "DatasheetUrl": datasheet_url,
        "ProductUrl": prod.get('ProductUrl', '').replace('.com', '.ca')
    }

    for param in prod.get('Parameters', []):
        name = param.get('ParameterText', '').strip().lower()
        if name: part_profile[name] = param.get('ValueText', '').strip()

    if datasheet_url:
        pdf_specs = auto_extract_specs_from_pdf(datasheet_url, list(part_profile.keys()))
        for k, v in pdf_specs.items():
            if k not in part_profile: part_profile[k] = v

    split_temperature_ranges(part_profile)
    return part_profile

def map_parameter_name(user_term, available_keys):
    user_term = user_term.lower().strip()
    if user_term in available_keys: return user_term
    for alias in SPEC_ALIASES.get(user_term, [user_term]):
        for key in available_keys:
            if alias in key: return key
    for key in available_keys:
        if user_term in key: return key
    return None

def find_similar_parts(token, search_keyword, constants_dict, variables_list):
    url = "https://api.digikey.com/products/v4/search/keyword"
    headers = {
        "Authorization": f"Bearer {token}", 
        "X-DIGIKEY-Client-Id": CLIENT_ID, 
        "Content-Type": "application/json",
        "X-DIGIKEY-Locale-Site": "US", 
        "X-DIGIKEY-Locale-Currency": "CAD"
    }
    response = requests.post(url, json={"Keywords": search_keyword, "Limit": 50}, headers=headers)
    
    if response.status_code != 200: 
        st.error(f"DigiKey API Error {response.status_code}: {response.text}")
        return None
        
    products = response.json().get('Products', [])
    if not products:
        st.error(f"DigiKey returned 0 parts when searching for '{search_keyword}'. Check for typos!")
        return []

    is_usd = (headers.get("X-DIGIKEY-Locale-Site") == "US")
    cad_exchange_rate = get_usd_to_cad_rate() if is_usd else 1.0

    matching_candidates = []
    for prod in products:
        cand_params = {p.get('ParameterText', '').strip().lower(): p.get('ValueText', '').strip() for p in prod.get('Parameters', []) if p.get('ParameterText')}
        split_temperature_ranges(cand_params)
        datasheet_url = prod.get('DatasheetUrl')
        pdf_scanned, match_failed = False, False

        keys_to_check = list(constants_dict.keys()) + variables_list
        is_missing_data = any(k not in cand_params or cand_params[k] == "-" for k in keys_to_check)
        
        if datasheet_url and is_missing_data:
            pdf_data = auto_extract_specs_from_pdf(datasheet_url, keys_to_check)
            for k, v in pdf_data.items():
                if k not in cand_params or cand_params[k] == "-": 
                    cand_params[k] = v

        for c_param, required_val in constants_dict.items():
            cand_val = cand_params.get(c_param)
            if not cand_val: match_failed = True; break
            
            req_str, cand_str = str(required_val).lower(), str(cand_val).lower()
            
            req_chunks = [x.strip() for x in req_str.replace('/', ',').split(',') if x.strip()]
            cand_chunks = [x.strip() for x in cand_str.replace('/', ',').split(',') if x.strip()]
            
            chunk_match = False
            for rc in req_chunks:
                for cc in cand_chunks:
                    if rc in cc or cc in rc:
                        chunk_match = True
                        break
                if chunk_match: break
                
            if chunk_match: continue
            
            req_nums = "".join(c for c in req_str if c.isdigit() or c == '.')
            cand_nums = "".join(c for c in cand_str if c.isdigit() or c == '.')
            if req_nums and cand_nums and (req_nums in cand_nums or cand_nums in req_nums): continue
            match_failed = True; break

        if match_failed: continue

        raw_price_1, raw_price_100 = extract_prices(prod)

        matching_candidates.append({
            "Part Number": prod.get('ManufacturerProductNumber') or prod.get('ProductCode'),
            "Manufacturer": prod.get('Manufacturer', {}).get('Name'),
            "Stock": prod.get('QuantityAvailable', 0),
            "Price1": raw_price_1 * cad_exchange_rate,
            "Price100": raw_price_100 * cad_exchange_rate,
            "DatasheetUrl": datasheet_url,
            "ProductUrl": prod.get('ProductUrl', '').replace('.com', '.ca'),
            "all_params": cand_params  
        })
    return matching_candidates

def generate_advanced_excel_buffer(ref_part, candidates, constants, variables):
    wb = Workbook()
    ws = wb.active
    center_align = Alignment(horizontal='center', vertical='center')
    
    start_params = 4
    num_params = len(constants) + len(variables)
    price1_col = start_params + num_params
    price100_col = price1_col + 1
    link_col = price100_col + 1

    ws.cell(row=1, column=1, value="Name / Part Number").alignment = center_align
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    ws.cell(row=1, column=2, value="Manufacturer").alignment = center_align
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    ws.cell(row=1, column=3, value="In Stock").alignment = center_align
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    
    if constants:
        ws.cell(row=1, column=start_params, value="Constant Aspects").alignment = center_align
        ws.merge_cells(start_row=1, start_column=start_params, end_row=1, end_column=start_params+len(constants)-1)
        
    if variables:
        ws.cell(row=1, column=start_params+len(constants), value="Varying Aspects").alignment = center_align
        ws.merge_cells(start_row=1, start_column=start_params+len(constants), end_row=1, end_column=start_params+num_params-1)
    
    ws.cell(row=1, column=price1_col, value="Price 1-Qty (CAD)").alignment = center_align
    ws.merge_cells(start_row=1, start_column=price1_col, end_row=2, end_column=price1_col)

    ws.cell(row=1, column=price100_col, value="Price 100-Qty (CAD)").alignment = center_align
    ws.merge_cells(start_row=1, start_column=price100_col, end_row=2, end_column=price100_col)

    ws.cell(row=1, column=link_col, value="DigiKey Link").alignment = center_align
    ws.merge_cells(start_row=1, start_column=link_col, end_row=2, end_column=link_col)

    col = start_params
    for c in constants + variables:
        ws.cell(row=2, column=col, value=str(c))
        col += 1

    ws.cell(row=3, column=1, value=f"ORIGINAL: {ref_part['Part Number']}")
    ws.cell(row=3, column=2, value=str(ref_part.get('Manufacturer', 'N/A')))
    ws.cell(row=3, column=3, value=ref_part.get('Stock', 0))
    col = start_params
    for c in constants + variables:
        ws.cell(row=3, column=col, value=str(ref_part.get(c, 'N/A')))
        col += 1
    
    ws.cell(row=3, column=price1_col, value=round(float(ref_part.get('Price1', 0.0)), 2))
    ws.cell(row=3, column=price1_col).number_format = '$0.00'
    ws.cell(row=3, column=price100_col, value=round(float(ref_part.get('Price100', 0.0)), 2))
    ws.cell(row=3, column=price100_col).number_format = '$0.00'
    
    url = ref_part.get('ProductUrl') or ref_part.get('DatasheetUrl')
    if url:
        ws.cell(row=3, column=link_col, value="View on DigiKey")
        ws.cell(row=3, column=link_col).hyperlink = url
        ws.cell(row=3, column=link_col).font = Font(color="0000FF", underline="single")
    else:
        ws.cell(row=3, column=link_col, value="N/A")

    current_row = 4
    for cand in candidates:
        ws.cell(row=current_row, column=1, value=str(cand['Part Number']))
        ws.cell(row=current_row, column=2, value=str(cand.get('Manufacturer', 'N/A')))
        ws.cell(row=current_row, column=3, value=cand.get('Stock', 0))
        col = start_params
        for c in constants + variables:
            val = cand['all_params'].get(c, 'N/A')
            ws.cell(row=current_row, column=col, value=str(val))
            if str(val) != str(ref_part.get(c, 'N/A')):
                ws.cell(row=current_row, column=col).font = Font(underline="single")
            col += 1
            
        ws.cell(row=current_row, column=price1_col, value=round(float(cand.get('Price1', 0.0)), 2))
        ws.cell(row=current_row, column=price1_col).number_format = '$0.00'
        ws.cell(row=current_row, column=price100_col, value=round(float(cand.get('Price100', 0.0)), 2))
        ws.cell(row=current_row, column=price100_col).number_format = '$0.00'

        if cand.get('ProductUrl'):
            ws.cell(row=current_row, column=link_col, value="View on DigiKey")
            ws.cell(row=current_row, column=link_col).hyperlink = cand.get('ProductUrl')
            ws.cell(row=current_row, column=link_col).font = Font(color="0000FF", underline="single")
        else:
            ws.cell(row=current_row, column=link_col, value="N/A")
            
        current_row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions[ws.cell(row=3, column=link_col).column_letter].width = 18
    for col_idx in range(start_params, price100_col + 1):
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = 20
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_batch_extract_excel_buffer(part_profiles, all_params):
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Extraction"
    bold_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Part Number").font = bold_font
    ws.cell(row=1, column=2, value="Manufacturer").font = bold_font
    ws.cell(row=1, column=3, value="In Stock").font = bold_font
    ws.cell(row=1, column=4, value="Price 1-Qty (CAD)").font = bold_font
    ws.cell(row=1, column=5, value="Price 100-Qty (CAD)").font = bold_font
    ws.cell(row=1, column=6, value="DigiKey Link").font = bold_font
    
    col = 7
    for param in all_params:
        ws.cell(row=1, column=col, value=str(param)).font = bold_font
        col += 1

    current_row = 2
    for profile in part_profiles:
        ws.cell(row=current_row, column=1, value=str(profile.get('Part Number', 'N/A')))
        ws.cell(row=current_row, column=2, value=str(profile.get('Manufacturer', 'N/A')))
        ws.cell(row=current_row, column=3, value=profile.get('Stock', 0))
        
        ws.cell(row=current_row, column=4, value=round(float(profile.get('Price1', 0.0)), 2))
        ws.cell(row=current_row, column=4).number_format = '$0.00'
        
        ws.cell(row=current_row, column=5, value=round(float(profile.get('Price100', 0.0)), 2))
        ws.cell(row=current_row, column=5).number_format = '$0.00'

        if profile.get('ProductUrl'):
            ws.cell(row=current_row, column=6, value="View on DigiKey")
            ws.cell(row=current_row, column=6).hyperlink = profile.get('ProductUrl')
            ws.cell(row=current_row, column=6).font = Font(color="0000FF", underline="single")
        else:
            ws.cell(row=current_row, column=6, value="N/A")

        col = 7
        for param in all_params:
            ws.cell(row=current_row, column=col, value=str(profile.get(param, 'N/A')))
            col += 1
        current_row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    for col_idx in range(7, 7 + len(all_params)):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ==========================================
# STREAMLIT UI
# ==========================================
st.title("Component Optimization Engine")
st.markdown("Easily find alternative electronic components and generate comparison spreadsheets.")

token = get_digikey_token()

if not token:
    st.stop()

tab1, tab2 = st.tabs(["Find Alternatives", "Batch Extract Characteristics"])

# --- TAB 1: FIND ALTERNATIVES ---
with tab1:
    st.markdown("Find alternative components based on strict rules.")
    ref_part_num = st.text_input("Enter Reference Part Number:", key="ref_input")
    
    if st.button("Lookup Part", type="primary"):
        if ref_part_num:
            with st.spinner("Fetching data..."):
                try:
                    profile = fetch_part_data(token, ref_part_num)
                    if profile:
                        st.session_state['ref_profile'] = profile
                    else:
                        st.error("Part not found on DigiKey.")
                except Exception as e:
                    st.error(str(e))
                    
    if 'ref_profile' in st.session_state:
        ref = st.session_state['ref_profile']
        st.success(f"Loaded: {ref['Part Number']} ({ref['Manufacturer']}) - {ref['Stock']} In Stock")
        
        keys = [k for k in ref.keys() if k not in ["Part Number", "Manufacturer", "Description", "Price1", "Price100", "Stock", "DatasheetUrl", "ProductUrl", "Category"]]
        
        def format_dropdown(key_name):
            return f"{key_name}: {ref.get(key_name, 'N/A')}"
            
        desc = str(ref.get("Description", ""))
        smart_sugg = " ".join(desc.split()[:2]) if desc else ""
        
        search_cat = st.text_input("Search Keyword (DigiKey shorthand):", value=smart_sugg)
        
        constants = st.multiselect("What MUST remain exactly the same?", options=keys, format_func=format_dropdown)
        
        for c in constants:
            val = str(ref.get(c, ""))
            if "~" in val or " to " in val.lower():
                st.warning(f"HEADS UP: You set '{c}' as a constant, but its value is a range ({val}). Try moving it to the 'can vary' list to avoid filtering out all parts!")

        variants = st.multiselect("What can vary?", options=[k for k in keys if k not in constants], format_func=format_dropdown)
            
        if st.button("Generate Comparison Spreadsheet"):
            if not search_cat:
                st.error("You must provide a search keyword!")
            else:
                with st.spinner("Scanning database and PDFs..."):
                    target_consts = {c: ref[c] for c in constants}
                    cands = find_similar_parts(token, search_cat, target_consts, variants)
                    
                    if cands is None:
                        pass 
                    elif not cands:
                        st.error("No matches found. Filters are too strict!")
                    else:
                        st.success(f"Found {len(cands)} matches!")
                        buffer = generate_advanced_excel_buffer(ref, cands, constants, variants)
                        
                        safe_filename = "".join(c for c in ref['Part Number'] if c.isalnum() or c in "-_")
                        st.download_button(
                            label="Download Comparison Spreadsheet",
                            data=buffer,
                            file_name=f"Comparison_{safe_filename}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

# --- TAB 2: BATCH EXTRACT ---
with tab2:
    st.markdown("Paste a list of part numbers (or PDF filenames) to extract all their characteristics into one master spreadsheet.")
    batch_input = st.text_area("Enter part numbers (comma-separated or one per line):", height=150)
    
    if st.button("Extract Data", type="primary"):
        if batch_input:
            parts = [p.strip() for p in batch_input.replace('\n', ',').split(',') if p.strip()]
            
            with st.spinner("Vacuuming data from DigiKey and PDFs..."):
                extracted = []
                all_params = set()
                
                for part in parts:
                    clean = part.replace('.pdf', '').replace('.PDF', '')
                    clean = re.sub(r'(?i)(^DS_|^infineon-|-datasheet-en|_en)', '', clean)
                    
                    try:
                        prof = fetch_part_data(token, clean)
                        if prof:
                            extracted.append(prof)
                            for k in prof.keys():
                                if k not in ["Part Number", "Manufacturer", "Description", "Price1", "Price100", "Stock", "DatasheetUrl", "ProductUrl", "Category"]:
                                    all_params.add(k)
                    except Exception:
                        pass
                
                if extracted:
                    st.success(f"Successfully processed {len(extracted)} parts!")
                    buffer = generate_batch_extract_excel_buffer(extracted, sorted(list(all_params)))
                    st.download_button(
                        label="Download Master Spreadsheet", 
                        data=buffer, 
                        file_name="Batch_Extraction.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("Failed to fetch any parts. Check your part numbers.")
